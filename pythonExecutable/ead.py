import openpyxl
from openpyxl.utils import get_column_letter
from lxml import etree
import tkinter as tk
from tkinter import filedialog
import os
import re
from datetime import datetime
import xml.dom.minidom as minidom

# Functions 

def convert_Date(inDate):
    if inDate.startswith("Jan"): return "01"
    if inDate.startswith("Feb"): return "02"
    if inDate.startswith("Mar"): return "03"
    if inDate.startswith("Apr"): return "04"
    if inDate.startswith("May"): return "05"
    if inDate.startswith("Jun"): return "06"
    if inDate.startswith("Jul"): return "07"
    if inDate.startswith("Aug"): return "08"
    if inDate.startswith("Sep"): return "09"
    if inDate.startswith("Oct"): return "10"
    if inDate.startswith("Nov"): return "11"
    if inDate.startswith("Dec"): return "12"

def endOfDecade(year):
    year = int(year)
    year += 9
    return year

def codedDate(i):
    
    # Undated 
    if i == 'undated':
        return '0000/0000'
    
    # 1 October-December, 2001
    elif re.search(r"([a-zA-Z]+).?\s*-\s*([a-zA-Z]+)\s*.?\s*(\d{4})",i):
        match = re.search(r"([a-zA-Z]+).?\s*-\s*([a-zA-Z]+)\s*.?\s*(\d{4})",i)
        year = match.group(3); 
        month = convert_Date(match.group(1)); 
        month2 = convert_Date(match.group(2))
        return str(year) + "-" + str(month) + "/" + str(year) + "-" + str(month2)
    # 2 January 24, 2014 - February 24, 2018 and a few variations Done
    elif re.search(r"([a-zA-Z]+)\s*,?\s*\b(\d{1,2})?(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\b\s*,?\s*(\d{4})?(\s*.{1,2}\b\s*([a-zA-Z]+)\s*,?\s*\b(\d{1,2})?(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\b\s*,?\s*(\d{4})?)",i) and not re.search("undated",i):
        match = re.search(r"([a-zA-Z]+)\s*,?\s*\b(\d{1,2})?(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\b\s*,?\s*(\d{4})?(\s*.{1,2}\b\s*([a-zA-Z]+)\s*,?\s*\b(\d{1,2})?(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\b\s*,?\s*(\d{4})?)",i)
        month = match.group(1); 
        month2 = match.group(5)
        if month:
            month = convert_Date(month); month = "-" + month
        if match.group(2):
            day = match.group(2)
            if len(day) < 2:
                day = day.insert(0, '0')
            day = "-" + day 
        year = match.group(3)
        if month2:
            month2 = convert_Date(month2); 
            month2 = "-" + month2
        if match.group(6):
            day2 = match.group(6)
            if len(day2) < 2:
                day2 = day2.insert(0, '0')
            day2 = "-" + day2 
        year2 = match.group(7)
        if re.search("(spring|summer|fall|winter)",i.lower()):
        #if (i like "*Spring*" or i like "*Fall*" or i like "*Summer*" or i like "*Winter*"):
            return str(year) + "/" + str(year2)
        elif not year:
            return str(year2) + str(month) + str(day) + "/" + str(year2) + str(month2) + str(day2)
        elif year2:
            return  str(year) + str(month) + str(day) + "/" + str(year2) + str(month2) + str(day2)
        else:
            return  str(year) + str(month) + str(day) + "/" + str(year) + str(month2) + str(day2)
    # 3 undated
    elif re.search("r(\d{4})?(?:-(\d{4}))?.*(?:\s*and\s*)?undated",i) and i not in "sfwxyswzFXSXyfqys":
        match = re.search("r(\d{4})?(?:-(\d{4}))?.*(?:\s*and\s*)?undated",i)
        year = None; year2 = None
        if match.group(1):
            year = match.group(1)
        if match.group(2):
            year2 = match.group(2)
            
        if year and year2:
            return str(year) + "/" + str(year2)
        else:
            return str(year)
    # 4 c 1790s, and 1790s
    elif re.search(r"^(c\.?\s+)?(\d{4})s$",i):
        match = re.search(r"^(c\.?\s+)?(\d{4})s$",i)
        year = match.group(2)
        year2 = endOfDecade(year)
        return str(year) + "/" + str(year2) 
    # 5 1970s-1980s
    elif re.search(r"^\s*(\d{4})s\s*-\s*(\d{4})s\s*$",i):
        match = re.search(r"^\s*(\d{4})s\s*-\s*(\d{4})s\s*$",i)
        year = match.group(1)
        year2 = match.group(2)
        year2 = endOfDecade(year2)
        return str(year) + "/" + str(year2)
    # 6 October, 2001
    elif re.search(r"^[a-zA-Z]+,?\s*(\d{4})$",i) and not re.search(r'(spring|summer|fall|winter)',i.lower()) and not re.search("^circa",i.lower()) :
        match = re.search(r"^[a-zA-Z]+,?\s*(\d{4})$",i)
        month = None; year = None
        if re.search(r"(^\w+)\b",i): 
            match = re.search(r"(^\w+)\b",i)
            month = convert_Date(match.group(1))       
        if re.search(r"(\d{4})",i): 
            match = re.search(r"(\d{4})",i)
            year = match.group(1)
        
        return str(year) + "-" + str(month)
    # 7 Spring, 2001
    elif re.search(r'spring|summer|fall|winter',i.lower()):
        match = re.search("(\d{4})$",i)
        if match.group(1):
            year = match.group(1)
        return str(year)
    # 8 October 16, 2001
    elif re.search(r"([a-zA-Z]+)\s*(\d{1,2})(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\s*,?\s*(\d{4})",i):
        match = re.search(r"([a-zA-Z]+)\s*(\d{1,2})(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\s*,?\s*(\d{4})",i)
        year = match.group(3); day = match.group(2); month = convert_Date(match.group(1)); 
        if len(day) < 2:
            day = day.insert(0, '0')
        return str(year) + "-" + str(month) + "-" + str(day)
    # 9 October 16-18, 2001
    elif re.search(r"([a-zA-Z]+)\s*(\d{1,2})(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\s*(?:.{1,2})\s*\b(\d{1,2})(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?,\s*(\d{4})",i) and i not in "hjnkejmnqwnmswdwfsvbkcfqelourpfvzsnfcgpsckwslrewhyozdhdsnafzojxez":
        match = re.search(r"([a-zA-Z]+)\s*(\d{1,2})(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?\s*(?:.{1,2})\s*\b(\d{1,2})(?:[nN][dD]|[sS][tT]|[rR][dD]|[tT][hH])?,\s*(\d{4})",i)
        year = match.group(4); day = match.group(2); day2 = match.group(3) ; month = convert_Date(match.group(1))
        if len(day) < 2:
            day = day.insert(0, '0')
        if len(day2) < 2:
            day2 = day2.insert(0, '0')
        return str(year) + "-" + str(month) + "-" + str(day) + "/" + str(year) + "-" + str(month) + "-" + str(day2) 
    # 10 c. 1945-1947
    elif re.search(r"^\s*c.\s*(\d{4})\s*-\s*(\d{4})\\s*$",i):
        match = re.search(r"^\s*c.\s*(\d{4})\s*-\s*(\d{4})\s*$",i)
        year = match.group(1); year2 = match.group(2)
        return str(year) + "/" + str(year2)
    # 11 1945 and c. 1945
    elif re.search(r"^\s*(?:c\.|[cC][iI][Rr][cC][aA].?)?\s*(\d{4})$", i):
        match = re.search(r"^\s*(?:c\.|[cC][iI][Rr][cC][aA].?)?\s*(\d{4})$", i)
        year = match.group(1)
        return str(year)
    # 13 1942, 1045, 1945-1947
    elif re.search(r"(\d.*\d)", i):
        match = re.search(r"(\d.*\d)", i)
        str2 = match.group(1)
        str2 = re.sub(r",\s|\s*-\s*", ",", str2)
        str3 = str2.split(",")
        year = min(map(int, str3))
        year2 = max(map(int, str3))
        return f"{year}/{year2}"


## Main Loop Function

def convert_to_xml(csv_file, xml):
    record = 1
    series_id = 1
    prev_c_num = 0
    element_stack = []
    global warnMsg 
    
    # Start Message
    print("Starting the script...", flush=True)
    
    for row in csv_file.iter_rows(min_row=2, values_only=True):
        
        # Set Vars
        v_series_id = str(row[0]).strip() if row[0] else None
        v_attribute = str(row[1]).strip() if row[1] else None
        v_c0 = int(row[2]) if row[2] else None
        v_box = int(row[3]) if row[3] else None
        v_file = int(row[4]) if row[4] else None
        v_title = str(row[5]).strip() if row[5] else None
        v_date = str(row[6]).strip() if row[6] else None
        v_dspace_url = str(row[8]).strip() if row[8] else None
        
        
        # Increase count of record to help identify errors.
        record += 1
        
        try:
            # Set a flag to determine if every cell is empty, blank, or contains only spaces
            all_cells_empty = True
                        
            # Loop through each property (cell) for the current row
            for property in row:
                # Check if the cell value is not null, not empty, and contains more than just spaces
                if property and str(property).strip():
                    all_cells_empty = False
                    break
            
            # If every cell is empty, blank or contains only spaces, skip the row
            if all_cells_empty:
                print(f"Warning: Blank row at Excel line: {record}", flush=True)
                continue
            
            # Data Checks - Errors and Warnings
            
            # Check for required information
            if not v_attribute or not v_c0 or not v_title:
                print(f"Error: Required record information missing for record at Excel line: {record}", flush=True)
                print("Press 'Enter' to exit...")
                input()
                exit()
                
            # Checks for High C#
            if v_c0 > 6:
                print(f"Warning: High c# - You may want to check your logic. - c# = {v_c0} at Excel line: {record}", flush=True)
                warnMsg = 1
            
            # Check for Series ID mismatch
            if v_series_id or (v_attribute == "series"):
                if not v_series_id or (re.sub("\D", "", v_series_id) != str(series_id)):
                    current_ser = "BLANK CELL" if not v_series_id or (not v_attribute) else v_series_id
                    print(f"Warning: Series ID mismatch for record at Excel line: {record} - ID in Record: {current_ser}, ID expected: ser{series_id}.", flush=True)
                    warnMsg = 1    
                series_id += 1
                
            
            # Current C# breaks ascending pattern.
            if v_c0 and (v_c0 > prev_c_num + 1):
                print(f"Warning: C# pattern broken on Excel line: {record}. Previous value: {prev_c_num}, Expecting value: {prev_c_num + 1}, actual value: {v_c0}.", flush=True)
                warnMsg = 1
            # Starting XML Building
                
            # Get the hierarchy level and inner text from the CSV row
            c_num = f"{v_c0:02d}" if v_c0 else None
            
            hierarchy = v_c0
            
            # Create a new cNum element
            new_element = xml.createElement(f"c{c_num}")
            
            # Create the 'did' element for new element.
            did = xml.createElement("did")
            new_element.appendChild(did)
            
            # Set Series ID
            if v_series_id:
                new_element.setAttribute("id", v_series_id) 
                
            # Set Level
            if v_attribute:
                new_element.setAttribute("level", v_attribute) 
            
            # Check if the 'Box' header exists
            if v_box:
                # Create Container Element.
                box = xml.createElement("container")
                # Add Container Inner Text
                box_text = str(v_box) if v_box else ""
                box.appendChild(xml.createTextNode(box_text))
                # Add Attribute
                box.setAttribute("type", "box") 
                did.appendChild(box) 
                
            # If not series or subseries populate empty value if no value given. 
            elif v_attribute not in ['subseries', 'series']:
                # Create Container Element.
                box = xml.createElement("container")
                # Add Container Inner Text
                box.appendChild(xml.createTextNode(""))
                # Add Attribute
                box.setAttribute("type", "box") 
                did.appendChild(box)
                
            # Check if the 'File' header exists
            if v_file:
                # Create Container Element.
                file = xml.createElement("container")
                # Add Container Inner Text
                file_text = str(v_file) if v_file else ""
                file.appendChild(xml.createTextNode(file_text))
                # Add Attribute
                file.setAttribute("type", "folder")  
                did.insertBefore(file, box)
                
            # If not series or subseries populate empty value if no value given. 
            elif v_attribute not in ['subseries', 'series']:
                # Create Container Element.
                file = xml.createElement("container")
                # Add Container Inner Text
                file.appendChild(xml.createTextNode(""))
                # Add Attribute
                file.setAttribute("type", "folder")  
                did.insertBefore(file, box)
                
            # Check if the 'Title' header exists
            if v_title:
                # Create the 'unittitle' child element of 'did' and set its inner text
                unittitle = xml.createElement("unittitle")
            
            # Check if 'extref' exists
            if v_dspace_url:
                # Create 'extref' element
                ext_ref = xml.createElement("extref")
                ext_ref.setAttribute("xmlns:xlink", "http://www.w3.org/1999/xlink")
                ext_ref.setAttribute("xlink:type", "simple")
                ext_ref.setAttribute("xlink:show", "new")
                ext_ref.setAttribute("xlink:actuate", "onRequest")
                ext_ref.setAttribute("xlink:href", v_dspace_url)

                # Set 'unittitle' text
                ext_ref_text = v_title if v_title else ""
                ext_ref.appendChild(xml.createTextNode(ext_ref_text))

                # Check if 'unitdate' exists
                if v_date:
                    # Create 'unitdate' element
                    unit_date = xml.createElement("unitdate")
                    unit_date.setAttribute("era", "ce")
                    unit_date.setAttribute("calendar", "gregorian")
                    #unit_date.setAttribute("normal", v_date)
                    unit_date.setAttribute("normal", codedDate(v_date))
                    unit_date.appendChild(xml.createTextNode(v_date))

                    # Append 'unitdate' to 'extref'
                    ext_ref.appendChild(unit_date)

                # Append 'extref' to 'unittitle'
                unittitle.appendChild(ext_ref)
            
            else:
                # Set 'unittitle' text
                unittitle_text = v_title if v_title else ""
                unittitle.appendChild(xml.createTextNode(unittitle_text))

                # Check if 'unitdate' exists
                if v_date:
                    # Create and append 'unitdate' element
                    unit_date = xml.createElement("unitdate")
                    unit_date.setAttribute("era", "ce")
                    unit_date.setAttribute("calendar", "gregorian")
                    #unit_date.setAttribute("normal", v_date)
                    unit_date.setAttribute("normal", codedDate(v_date))
                    unit_date.appendChild(xml.createTextNode(v_date))
                    unittitle.appendChild(unit_date)
            
            did.appendChild(unittitle) 
            
            # Handle the hierarchy
            if len(element_stack) == 0:
                # If the stack is empty, add the element as a child of the root
                rootElement.appendChild(new_element) 
                
            elif hierarchy and (hierarchy < int(element_stack[-1].nodeName[1:])):
                # If the hierarchy is less than the current open element, close the open element and add the new element
                while hierarchy < int(element_stack[-1].nodeName[1:]):
                    element_stack.pop()
                    
                element_stack[-1].appendChild(new_element)
                
            elif hierarchy and (hierarchy == int(element_stack[-1].nodeName[1:])):
                # If the hierarchy is equal to the current open element, add the new element as a sibling
                # append to parent of last element in stack
                element_stack[-1].parentNode.appendChild(new_element)
            elif hierarchy and (hierarchy > int(element_stack[-1].nodeName[1:])):
                # If the hierarchy is greater than the current open element, add the new element as a child of the current open element
                element_stack[-1].appendChild(new_element)
                
            # Add the new element to the stack of open elements
            element_stack.append(new_element)
            
            # Set the previous c# to the current to use in comparison in the next iteration
            prev_c_num = v_c0
        
        except BaseException as e:
            print(str(e))
            print(f"Error: Could not process record at Excel line: {record}", flush=True)
            input()
            exit()


# Vars
warnMsg = None

# Set the current directory as the starting location for the file picker
root = tk.Tk()
root.withdraw()
excel_file_path = filedialog.askopenfilename(initialdir=os.getcwd(), filetypes=[("Excel Files", "*.xlsx")])

# Load the workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Get the desired sheet (or the first sheet if "template" doesn't exist)
if "template" in workbook.sheetnames:
    sheet = workbook["template"]
else:
    sheet = workbook.active

# Create the XML document
# Example of usage
xml_doc = minidom.Document()
rootElement = xml_doc.createElement("RootElement")
xml_doc.appendChild(rootElement)

# Convert Excel to XML 
convert_to_xml(sheet, xml_doc)

# Save the XML document
filepath = os.getcwd()
fileName = "output_file.xml"
fullpath = os.path.join(filepath, fileName)

with open(fullpath, 'w') as f:
    f.write(xml_doc.toprettyxml(indent="  "))   


# Stop message
print(f"Script completed. Results written to: {fullpath}", end="", flush=True) 
print("\033[32m") # ANSI Escape code for setting console text color to green

# Pause at the end if warnings happened during run. 
if warnMsg:
    print("Warnings occoured during run.")
    input("Press 'Enter' to exit and open the output file...")

# Open Saved file in notepad
# open saved file using notepad.exe

# Open saved xml file remove the top and bottom two lines, then save it again.
# set the file name and open the file

with open(fullpath, "r") as file:
    # read the content of the file
    content = file.readlines()

# remove the top two lines and bottom two lines of the file
content = content[2:-1]

# save the modified content to the same file
with open(fullpath, "w") as file:
    file.writelines(content)


os.startfile(fullpath)
